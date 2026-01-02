import pandas as pd
import numpy as np
from datetime import datetime, timedelta
import uvicorn
from fastapi import FastAPI, Request, HTTPException, Cookie
from fastapi.responses import HTMLResponse, JSONResponse, StreamingResponse
import os
import socket
import secrets
from PIL import Image, ImageDraw, ImageFont
import io
import base64
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter

# ==================== WARRANTY DATA STORAGE ====================

WARRANTY_DATA = {
    "credit_df": None,
    "debit_df": None,
    "arbitration_df": None,  # NOW: Pending Arbitration month-wise summary
    "source_df": None,
    "current_month_df": None,
    "current_month_source_df": None,
    "compensation_df": None,
    "compensation_source_df": None,
    "pr_approval_df": None,
    "pr_approval_source_df": None,
}

# ==================== FILE HELPERS ====================

def find_data_file(filename: str):
    """Find data file in multiple possible locations"""
    possible_paths = [
        f"/mnt/data/{filename}",
        filename,
        f"./{filename}",
        f"Data/{filename}",
        f"data/{filename}",
    ]
    for path in possible_paths:
        if os.path.exists(path):
            print(f"  Found: {filename} at {path}")
            return path
    print(f"  WARNING: {filename} not found. Checked: {possible_paths}")
    return None

# ==================== PR APPROVAL ====================

def process_pr_approval():
    input_path = find_data_file("Pr_Approval_Claims_Merged.xlsx")
    if input_path is None:
        print("  PR Approval file not found - returning empty data")
        return None, None

    try:
        df = pd.read_excel(input_path)
        print("  PR Approval data loaded successfully")
        print(f"  Total rows in source data: {len(df)}")

        summary_columns = [
            "Division",
            "PA Request No.",
            "PA Date",
            "Request Type",
            "App. Claim Amt from M&M",
        ]
        available_summary_columns = [c for c in summary_columns if c in df.columns]
        if not available_summary_columns:
            print("  No required columns found in PR Approval file")
            return None, None

        df_display = df[available_summary_columns].copy()

        if "Division" in df_display.columns:
            df_display["Division"] = df_display["Division"].astype(str).str.strip()
            df_display = df_display[
                df_display["Division"].notna()
                & (df_display["Division"] != "")
                & (df_display["Division"] != "nan")
            ]

        if "App. Claim Amt from M&M" in df_display.columns:
            df_display["App. Claim Amt from M&M"] = pd.to_numeric(
                df_display["App. Claim Amt from M&M"], errors="coerce"
            ).fillna(0)

        summary_data = []
        if "Division" in df_display.columns:
            for division in sorted(df_display["Division"].unique()):
                div_data = df_display[df_display["Division"] == division]
                row = {"Division": division}
                row["Total Requests"] = len(div_data)
                if "App. Claim Amt from M&M" in df_display.columns:
                    row["Total Approved Amount"] = div_data["App. Claim Amt from M&M"].sum()
                if "Request Type" in df_display.columns:
                    request_types = div_data["Request Type"].value_counts().to_dict()
                    for k, v in request_types.items():
                        if pd.notna(k) and str(k).strip() != "":
                            row[f"{k} Count"] = v
                summary_data.append(row)

            summary_df = pd.DataFrame(summary_data)

            grand_total = {"Division": "Grand Total"}
            for col in summary_df.columns:
                if col != "Division" and summary_df[col].dtype in ["int64", "float64"]:
                    grand_total[col] = summary_df[col].sum()
            summary_df = pd.concat([summary_df, pd.DataFrame([grand_total])], ignore_index=True)
        else:
            summary_df = pd.DataFrame()

        return summary_df, df

    except Exception as e:
        import traceback
        print(f"  Error processing PR Approval data: {e}")
        traceback.print_exc()
        return None, None

# ==================== COMPENSATION CLAIM ====================

def process_compensation_claim():
    input_path = find_data_file("Transit_Claims_Merged.xlsx")
    if input_path is None:
        print("  Compensation Claim file not found - returning empty data")
        return None, None

    try:
        df = pd.read_excel(input_path)
        print("  Compensation Claim data loaded successfully")
        print(f"  Total rows in source data: {len(df)}")

        required_columns = [
            "Division",
            "RO Id.",
            "Registration No.",
            "RO Date",
            "RO Bill Date",
            "Chassis No.",
            "Model Group",
            "Claim Amount",
            "Request Status",
            "Claim Approved Amt.",
            "No. of Days",
        ]
        available_columns = [c for c in required_columns if c in df.columns]
        if not available_columns:
            print("  No required columns found in Compensation Claim file")
            return None, None

        df_filtered = df[available_columns].copy()

        if "Division" in df_filtered.columns:
            df_filtered["Division"] = df_filtered["Division"].astype(str).str.strip()
            df_filtered = df_filtered[
                df_filtered["Division"].notna()
                & (df_filtered["Division"] != "")
                & (df_filtered["Division"] != "nan")
            ]

        if "RO Id." in df_filtered.columns:
            def format_ro_id(x):
                if pd.isna(x) or str(x).strip() == "":
                    return ""
                try:
                    return f"RO{str(int(float(x)))}"
                except Exception:
                    s = str(x).strip()
                    return s if s.startswith("RO") else f"RO{s}"
            df_filtered["RO Id."] = df_filtered["RO Id."].apply(format_ro_id)

        for col in ["Claim Amount", "Claim Approved Amt.", "No. of Days"]:
            if col in df_filtered.columns:
                df_filtered[col] = pd.to_numeric(df_filtered[col], errors="coerce").fillna(0)

        summary_data = []
        if "Division" in df_filtered.columns:
            for division in sorted(df_filtered["Division"].unique()):
                div_data = df_filtered[df_filtered["Division"] == division]
                row = {"Division": division}
                row["Total Claims"] = len(div_data)
                if "Claim Amount" in df_filtered.columns:
                    row["Total Claim Amount"] = div_data["Claim Amount"].sum()
                if "Claim Approved Amt." in df_filtered.columns:
                    row["Total Approved Amount"] = div_data["Claim Approved Amt."].sum()
                if "No. of Days" in df_filtered.columns:
                    row["Avg No. of Days"] = div_data["No. of Days"].mean()
                summary_data.append(row)

            summary_df = pd.DataFrame(summary_data)
            grand_total = {"Division": "Grand Total"}
            if "Total Claims" in summary_df.columns:
                grand_total["Total Claims"] = summary_df["Total Claims"].sum()
            if "Total Claim Amount" in summary_df.columns:
                grand_total["Total Claim Amount"] = summary_df["Total Claim Amount"].sum()
            if "Total Approved Amount" in summary_df.columns:
                grand_total["Total Approved Amount"] = summary_df["Total Approved Amount"].sum()
            if "Avg No. of Days" in summary_df.columns:
                grand_total["Avg No. of Days"] = summary_df["Avg No. of Days"].mean()

            summary_df = pd.concat([summary_df, pd.DataFrame([grand_total])], ignore_index=True)
        else:
            summary_df = pd.DataFrame()

        return summary_df, df_filtered

    except Exception as e:
        import traceback
        print(f"  Error processing compensation claim data: {e}")
        traceback.print_exc()
        return None, None

# ==================== CURRENT MONTH WARRANTY ====================

def process_current_month_warranty():
    input_path = find_data_file("Pending Warranty Claim Details.xlsx")
    if input_path is None:
        print("  Current Month Warranty file not found - returning empty data")
        return None, None

    try:
        df = pd.read_excel(input_path, sheet_name="Pending Warranty Claim Details")
        print("  Current Month Warranty data loaded successfully")
        print(f"  Total rows in source data: {len(df)}")

        required_columns = ["Division", "Pending Claims Spares", "Pending Claims Labour"]
        missing = [c for c in required_columns if c not in df.columns]
        if missing:
            print(f"  Missing columns in Current Month Warranty: {missing}")
            return None, None

        df["Division"] = df["Division"].astype(str).str.strip()
        df = df[df["Division"].notna() & (df["Division"] != "") & (df["Division"] != "nan")]

        summary_data = []
        for division in sorted(df["Division"].unique()):
            div_data = df[df["Division"] == division]
            spares_count = div_data["Pending Claims Spares"].notna().sum()
            labour_count = div_data["Pending Claims Labour"].notna().sum()
            summary_data.append(
                {
                    "Division": division,
                    "Pending Claims Spares Count": spares_count,
                    "Pending Claims Labour Count": labour_count,
                    "Total Pending Claims": spares_count + labour_count,
                }
            )

        summary_df = pd.DataFrame(summary_data)
        grand_total = {
            "Division": "Grand Total",
            "Pending Claims Spares Count": summary_df["Pending Claims Spares Count"].sum(),
            "Pending Claims Labour Count": summary_df["Pending Claims Labour Count"].sum(),
            "Total Pending Claims": summary_df["Total Pending Claims"].sum(),
        }
        summary_df = pd.concat([summary_df, pd.DataFrame([grand_total])], ignore_index=True)
        return summary_df, df

    except Exception as e:
        import traceback
        print(f"  Error processing current month warranty data: {e}")
        traceback.print_exc()
        return None, None

# ==================== WARRANTY CREDIT/DEBIT + PENDING ARB MONTHWISE ====================

def process_warranty_data():
    input_path = find_data_file("Warranty Debit.xlsx")
    if input_path is None:
        print("  Warranty Debit file not found - returning empty data")
        return None, None, None, None

    try:
        df = pd.read_excel(input_path, sheet_name="Sheet1")
        print("  Warranty data loaded successfully")
        print(f"  Total rows in source data: {len(df)}")

        dealer_mapping = {
            "AMRAVATI": "AMT",
            "CHAUFULA_SZZ": "CHA",
            "CHIKHALI": "CHI",
            "KOLHAPUR_WS": "KOL",
            "NAGPUR_KAMPTHEE ROAD": "HO",
            "NAGPUR_WARDHAMAN NGR": "CITY",
            "SHIKRAPUR_SZS": "SHI",
            "WAGHOLI": "WAG",
            "YAVATMAL": "YAT",
            "NAGPUR_WARDHAMAN NGR_CQ": "CQ",
        }

        # Clean numeric columns
        numeric_columns = ["Total Claim Amount", "Credit Note Amount", "Debit Note Amount"]
        for col in numeric_columns:
            if col in df.columns:
                df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0)
            else:
                df[col] = 0

        df["Dealer_Code"] = df["Dealer Location"].map(dealer_mapping).fillna(df["Dealer Location"])
        df["Month"] = df["Fiscal Month"].astype(str).str.strip().str[:3]
        df["Claim arbitration ID"] = df["Claim arbitration ID"].astype(str).replace("nan", "").replace("", np.nan)

        dealers = sorted(df["Dealer_Code"].unique())
        months = ["Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]

        # CREDIT SUMMARY
        credit_df = pd.DataFrame({"Division": dealers})
        for month in months:
            month_data = df[df["Month"] == month]
            if not month_data.empty:
                summary = month_data.groupby("Dealer_Code")["Credit Note Amount"].sum().reset_index()
                summary.columns = ["Division", f"Credit Note {month}"]
                credit_df = credit_df.merge(summary, on="Division", how="left")
            else:
                credit_df[f"Credit Note {month}"] = 0
        credit_df = credit_df.fillna(0)
        credit_cols = [f"Credit Note {m}" for m in months]
        credit_df["Total Credit"] = credit_df[credit_cols].sum(axis=1)
        grand_total_credit = {"Division": "Grand Total"}
        for col in credit_df.columns[1:]:
            grand_total_credit[col] = credit_df[col].sum()
        credit_df = pd.concat([credit_df, pd.DataFrame([grand_total_credit])], ignore_index=True)

        # DEBIT SUMMARY
        debit_df = pd.DataFrame({"Division": dealers})
        for month in months:
            month_data = df[df["Month"] == month]
            if not month_data.empty:
                summary = month_data.groupby("Dealer_Code")["Debit Note Amount"].sum().reset_index()
                summary.columns = ["Division", f"Debit Note {month}"]
                debit_df = debit_df.merge(summary, on="Division", how="left")
            else:
                debit_df[f"Debit Note {month}"] = 0
        debit_df = debit_df.fillna(0)
        debit_cols = [f"Debit Note {m}" for m in months]
        debit_df["Total Debit"] = debit_df[debit_cols].sum(axis=1)
        grand_total_debit = {"Division": "Grand Total"}
        for col in debit_df.columns[1:]:
            grand_total_debit[col] = debit_df[col].sum()
        debit_df = pd.concat([debit_df, pd.DataFrame([grand_total_debit])], ignore_index=True)

        # CLAIM ARBITRATION TAB SHOULD SHOW: PENDING ARBITRATION MONTHWISE + TOTAL PENDING
        arbitration_df = pd.DataFrame({"Division": dealers})
        print("\n  Processing Pending Arbitration (Month-wise)...")

        def is_empty_or_hyphen(value):
            if pd.isna(value):
                return True
            v = str(value).strip()
            if v == "" or v == "-" or v.upper() == "NAN":
                return True
            return False

        for month in months:
            month_data = df[df["Month"] == month].copy()
            month_data["Is_Pending_ARB"] = month_data["Claim arbitration ID"].apply(is_empty_or_hyphen)
            month_data["Pending_Arb_Amount"] = month_data.apply(
                lambda r: r["Debit Note Amount"] if r["Is_Pending_ARB"] else 0,
                axis=1,
            )

            pend_summary = month_data.groupby("Dealer_Code")["Pending_Arb_Amount"].sum().reset_index()
            pend_summary.columns = ["Division", f"Pending Arbitration {month}"]
            arbitration_df = arbitration_df.merge(pend_summary, on="Division", how="left")

        arbitration_df = arbitration_df.fillna(0)
        pending_cols = [f"Pending Arbitration {m}" for m in months]
        arbitration_df["Total Pending Arbitration"] = arbitration_df[pending_cols].sum(axis=1)

        grand_total_pend = {"Division": "Grand Total"}
        for col in arbitration_df.columns[1:]:
            grand_total_pend[col] = arbitration_df[col].sum()
        arbitration_df = pd.concat([arbitration_df, pd.DataFrame([grand_total_pend])], ignore_index=True)

        return credit_df, debit_df, arbitration_df, df

    except Exception as e:
        import traceback
        print(f"  Error processing warranty data: {e}")
        traceback.print_exc()
        return None, None, None, None

# ==================== IMAGE HANDLING ====================

def get_mahindra_images():
    """Load Mahindra vehicle images from the folder"""
    image_folder = r"D:\Power BI New\Warranty Debit\Image"
    images = []
    branding_images = []
    vehicle_images = []

    if os.path.exists(image_folder):
        try:
            for file in os.listdir(image_folder):
                if file.lower().endswith((".png", ".jpg", ".jpeg", ".gif", ".bmp")):
                    image_path = os.path.join(image_folder, file)
                    try:
                        with open(image_path, "rb") as img_file:
                            img_data = base64.b64encode(img_file.read()).decode()
                            img_dict = {"name": file, "data": img_data, "path": image_path}
                            file_lower = file.lower()
                            if "mahindra" in file_lower or "logo" in file_lower or "hero" in file_lower:
                                branding_images.append(img_dict)
                            else:
                                vehicle_images.append(img_dict)
                    except Exception as e:
                        print(f"  Could not load {file}: {e}")
        except Exception as e:
            print(f"  Error reading image folder: {e}")
    else:
        print(f"  Image folder not found: {image_folder}")

    images = branding_images + vehicle_images
    return images

print("Loading Mahindra vehicle images...")
MAHINDRA_IMAGES = get_mahindra_images()
print(f"Loaded {len(MAHINDRA_IMAGES)} images")

# ==================== AUTH ====================

USER_CREDENTIALS = {}
SESSIONS = {}

def load_user_credentials():
    """Load user credentials from UserID.xlsx (kept as-is for login)"""
    try:
        user_file = r"D:\Power BI New\Warranty Debit\UserID.xlsx"
        if not os.path.exists(user_file):
            print(f"  ERROR: User file not found: {user_file}")
            return {}

        df = pd.read_excel(user_file)
        credentials = {}
        for _, row in df.iterrows():
            try:
                uid_value = row.get("User ID", None)
                pwd_value = row.get("Password", None)
                if pd.isna(uid_value) or pd.isna(pwd_value):
                    continue
                user_id = str(int(float(uid_value)))
                password = str(pwd_value).strip()
                if user_id and password:
                    credentials[user_id] = password
            except Exception:
                continue
        return credentials
    except Exception as e:
        print(f"  ERROR loading credentials: {e}")
        return {}

USER_CREDENTIALS = load_user_credentials()

class CaptchaGenerator:
    @staticmethod
    def generate_captcha(length=6):
        allowed_chars = "ABCDEFGHJKLMNPQRSTUVWXYZ123456789"
        captcha_text = "".join(secrets.choice(allowed_chars) for _ in range(length))

        width, height = 500, 150
        image = Image.new("RGB", (width, height), color="white")
        draw = ImageDraw.Draw(image)

        for _ in range(5):
            x1 = secrets.randbelow(width)
            y1 = secrets.randbelow(height)
            x2 = secrets.randbelow(width)
            y2 = secrets.randbelow(height)
            draw.line((x1, y1, x2, y2), fill="lightgray", width=1)

        try:
            font = ImageFont.truetype("C:\\Windows\\Fonts\\arial.ttf", 80)
        except Exception:
            try:
                font = ImageFont.truetype("/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf", 80)
            except Exception:
                font = ImageFont.load_default()

        x_offset = 15
        for i, ch in enumerate(captcha_text):
            y_offset = np.random.randint(15, 50)
            draw.text((x_offset + i * 70, y_offset), ch, fill="#FF8C00", font=font)

        for _ in range(50):
            x = secrets.randbelow(width)
            y = secrets.randbelow(height)
            draw.point((x, y), fill="#FFD699")

        img_io = io.BytesIO()
        image.save(img_io, "PNG")
        img_io.seek(0)
        img_base64 = base64.b64encode(img_io.getvalue()).decode()
        return captcha_text, f"data:image/png;base64,{img_base64}"

def create_session(user_id: str):
    session_id = secrets.token_hex(16)
    SESSIONS[session_id] = {
        "user_id": user_id,
        "created_at": datetime.now(),
        "last_activity": datetime.now(),
    }
    return session_id

def verify_session(session_id: str):
    if session_id not in SESSIONS:
        return None
    session = SESSIONS[session_id]
    if (datetime.now() - session["last_activity"]).total_seconds() > 8 * 3600:
        del SESSIONS[session_id]
        return None
    session["last_activity"] = datetime.now()
    return session["user_id"]

# ==================== LOGIN PAGE (UNCHANGED UI) ====================

LOGIN_PAGE = """
<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Warranty Management System - Login</title>
    <style>
        * { margin:0; padding:0; box-sizing:border-box; }
        body{
            font-family:'Segoe UI', Tahoma, Geneva, Verdana, sans-serif;
            background: linear-gradient(135deg, #FF8C00 0%, #FF6B35 100%);
            height:100vh; width:100vw;
            display:flex; justify-content:center; align-items:center;
            overflow:hidden;
        }
        .login-wrapper{
            background:white; border-radius:20px;
            box-shadow:0 20px 60px rgba(0,0,0,0.3);
            overflow:hidden; width:95vw; height:95vh;
            max-width:1400px; max-height:800px;
            display:grid; grid-template-columns:1fr 1fr;
        }
        .login-left{ padding:30px; display:flex; flex-direction:column; justify-content:center; }
        .login-right{
            background: linear-gradient(135deg, #FF8C00 0%, #FF6B35 100%);
            padding:25px; display:flex; justify-content:center; align-items:center; color:white;
        }
        .logo-section{ text-align:center; margin-bottom:15px; }
        .logo-section h1{ font-size:28px; color:#333; margin-bottom:8px; }
        .login-form{ display:flex; flex-direction:column; gap:20px; }
        .form-group{ display:flex; flex-direction:column; gap:8px; }
        .form-group label{ font-weight:600; color:#333; font-size:14px; }
        .form-group input{
            padding:12px; border:2px solid #e0e0e0; border-radius:8px; font-size:14px;
        }
        .form-group input:focus{ outline:none; border-color:#FF8C00; box-shadow:0 0 8px rgba(255,140,0,0.2); }
        .login-btn{
            background: linear-gradient(135deg, #FF8C00 0%, #FF6B35 100%);
            color:white; border:none; padding:12px; border-radius:8px; font-weight:600;
            cursor:pointer; margin-top:10px;
        }
        .error-message{ color:#c62828; font-size:13px; margin-top:-15px; display:none; }
        .error-message.show{ display:block; }
        .captcha-section{ margin-top:15px; padding:10px; background:#f5f5f5; border-radius:8px; }
        .captcha-image{ width:100%; height:auto; margin-bottom:10px; border-radius:4px; }
        .right-content{ text-align:center; }
        .right-content h2{ font-size:32px; margin-bottom:20px; }
        .right-content p{ font-size:16px; line-height:1.6; opacity:0.95; }
    </style>
</head>
<body>
    <div class="login-wrapper">
        <div class="login-left">
            <div class="logo-section">
                <h1>Unnati Motors Warranty Management</h1>
                <p style="color:#666; font-size:14px;">Mahindra All Division Warranty Overview Dashboard</p>
                <p style="color:#666; font-size:14px;">Enter your credentials to access the warranty dashboard</p>
            </div>
            <form class="login-form" onsubmit="handleLogin(event)">
                <div class="form-group">
                    <label for="userId">User ID</label>
                    <input type="text" id="userId" name="userId" placeholder="Enter your User ID" required>
                </div>
                <div class="form-group">
                    <label for="password">Password</label>
                    <input type="password" id="password" name="password" placeholder="Enter your password" required>
                </div>
                <div class="captcha-section">
                    <img id="captchaImage" class="captcha-image" src="" alt="CAPTCHA">
                    <input type="text" id="captchaInput" placeholder="Enter CAPTCHA" required style="width:100%; padding:8px; border:2px solid #e0e0e0; border-radius:4px;">
                </div>
                <div class="error-message" id="errorMessage"></div>
                <button type="submit" class="login-btn">Login</button>
            </form>
        </div>
        <div class="login-right">
            <div class="right-content">
                <h2>Welcome</h2>
                <p>Welcome to Warranty Management System</p>
            </div>
        </div>
    </div>

    <script>
        let currentCaptcha = '';

        async function loadCaptcha(){
            try{
                const response = await fetch('/api/captcha');
                const data = await response.json();
                currentCaptcha = data.captcha;
                document.getElementById('captchaImage').src = data.image;
            }catch(e){
                console.error('Error loading CAPTCHA:', e);
            }
        }

        async function handleLogin(event){
            event.preventDefault();

            const userId = document.getElementById('userId').value;
            const password = document.getElementById('password').value;
            const captchaInput = document.getElementById('captchaInput').value;
            const errorDiv = document.getElementById('errorMessage');

            if(captchaInput.toUpperCase() !== currentCaptcha){
                errorDiv.textContent = 'CAPTCHA is incorrect';
                errorDiv.classList.add('show');
                loadCaptcha();
                return;
            }

            try{
                const response = await fetch('/api/login', {
                    method:'POST',
                    headers:{ 'Content-Type':'application/json' },
                    body: JSON.stringify({ user_id:userId, password:password })
                });

                if(response.ok){
                    window.location.href = '/dashboard';
                }else{
                    const err = await response.json();
                    errorDiv.textContent = (err.detail || 'Login failed');
                    errorDiv.classList.add('show');
                    loadCaptcha();
                }
            }catch(e){
                errorDiv.textContent = 'Error: ' + e.message;
                errorDiv.classList.add('show');
            }
        }

        window.onload = function(){ loadCaptcha(); };
    </script>
</body>
</html>
"""

# ==================== DASHBOARD HTML (REMOVED WELCOME/CHANGE PASSWORD/LOGOUT) ====================

DASHBOARD_HTML = """
<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Unnati Warranty Management Dashboard</title>
    <style>
        * { margin: 0; padding: 0; box-sizing: border-box; }
        body{
            font-family:'Segoe UI', Tahoma, Geneva, Verdana, sans-serif;
            background: linear-gradient(135deg, #f5f5f5 0%, #e0e0e0 100%);
            min-height:100vh;
        }
        .navbar{
            background: linear-gradient(135deg, #FF8C00 0%, #FF6B35 100%);
            color:white; padding:15px 0;
            box-shadow:0 2px 8px rgba(0,0,0,0.15);
            position:sticky; top:0; z-index:100;
        }
        .navbar .container-fluid{
            max-width:1400px; margin:0 auto;
            display:flex; justify-content:center; align-items:center;
            padding:0 30px;
        }
        .navbar-brand{ font-size:24px; font-weight:700; }
        .container{ max-width:1400px; margin:30px auto; padding:0 20px; }
        .dashboard-content{
            background:white; border-radius:12px;
            box-shadow:0 2px 10px rgba(0,0,0,0.1);
            padding:30px;
        }
        .nav-tabs{ border-bottom:2px solid #FF8C00; margin-bottom:30px; }
        .nav-tabs .nav-link{
            color:#666; font-weight:600; border:none;
            border-bottom:3px solid transparent;
            padding:12px 20px; cursor:pointer;
        }
        .nav-tabs .nav-link:hover{ color:#FF8C00; border-bottom-color:#FF8C00; }
        .nav-tabs .nav-link.active{ color:#FF8C00; border-bottom-color:#FF8C00; background:transparent; }
        .tab-content{ display:none; }
        .tab-content.active{ display:block; }
        .data-table{
            width:100%; border-collapse:collapse;
            margin-top:20px; font-size:12px;
        }
        .data-table thead th{
            background: linear-gradient(135deg, #FF8C00 0%, #FF6B35 100%);
            color:white; padding:12px; text-align:center;
            font-weight:600; border:none; font-size:11px;
        }
        .data-table tbody td{
            padding:10px 12px; border-bottom:1px solid #e0e0e0;
            text-align:right;
        }
        .data-table tbody td:first-child{
            text-align:left; font-weight:600; color:#333;
        }
        .data-table tbody tr:hover{ background:#f9f9f9; }
        .data-table tbody tr:last-child{
            background:#fff8f3; font-weight:700;
            border-top:2px solid #FF8C00; border-bottom:2px solid #FF8C00;
        }
        .data-table tbody tr:last-child td{ color:#FF8C00; }
        .loading-spinner{ display:none; text-align:center; padding:40px; }
        .spinner{
            border:4px solid rgba(255,140,0,0.2);
            border-top:4px solid #FF8C00;
            border-radius:50%; width:40px; height:40px;
            animation:spin 1s linear infinite; margin:0 auto;
        }
        @keyframes spin{ 0%{transform:rotate(0deg);} 100%{transform:rotate(360deg);} }
        .table-title{ font-size:16px; font-weight:700; color:#FF8C00; margin-bottom:15px; }
        .table-wrapper{ overflow-x:auto; }

        .export-section{
            margin:30px 0; padding:20px;
            background: linear-gradient(135deg, #fff8f3 0%, #ffe8d6 100%);
            border-radius:8px; border-left:5px solid #FF8C00;
            box-shadow:0 2px 8px rgba(255,140,0,0.1);
        }
        .export-section h3{ color:#FF8C00; margin-bottom:15px; font-size:16px; font-weight:700; }
        .export-controls{
            display:flex; gap:15px; align-items:center; flex-wrap:wrap;
            background:white; padding:15px; border-radius:6px;
        }
        .export-control-group{ display:flex; gap:8px; align-items:center; }
        .export-control-group label{ font-weight:600; color:#333; font-size:14px; min-width:80px; }
        .export-control-group select{
            padding:8px 12px; border:2px solid #FF8C00; border-radius:4px;
            cursor:pointer; background:white; font-size:13px; min-width:150px;
        }
        .export-btn{
            padding:10px 25px;
            background: linear-gradient(135deg, #4CAF50 0%, #45a049 100%);
            color:white; border:none; border-radius:4px;
            cursor:pointer; font-weight:700; font-size:14px;
        }
        .export-btn:disabled{ background:#ccc; cursor:not-allowed; }
    </style>
</head>
<body>
    <nav class="navbar navbar-dark">
        <div class="container-fluid">
            <span class="navbar-brand">Unnati Motors Warranty Management Dashboard</span>
        </div>
    </nav>

    <div class="container">
        <div class="dashboard-content">
            <div class="loading-spinner" id="loadingSpinner">
                <div class="spinner"></div>
                <p style="margin-top:15px; color:#666;">Loading warranty data...</p>
            </div>

            <div id="warrantyTabs" style="display:none;">
                <div class="nav-tabs">
                    <button class="nav-link active" onclick="switchTab('credit')">Warranty Credit</button>
                    <button class="nav-link" onclick="switchTab('debit')">Warranty Debit</button>
                    <button class="nav-link" onclick="switchTab('arbitration')">Claim Arbitration</button>
                    <button class="nav-link" onclick="switchTab('currentmonth')">Current Month Warranty</button>
                    <button class="nav-link" onclick="switchTab('compensation')">Compensation Claim</button>
                    <button class="nav-link" onclick="switchTab('pr_approval')">PR Approval</button>
                </div>

                <div class="export-section">
                    <h3>Export to Excel</h3>
                    <div class="export-controls">
                        <div class="export-control-group">
                            <label for="divisionFilter">Division:</label>
                            <select id="divisionFilter">
                                <option value="">-- Select Division --</option>
                                <option value="All">All Divisions</option>
                            </select>
                        </div>

                        <div class="export-control-group">
                            <label for="exportType">Export Type:</label>
                            <select id="exportType">
                                <option value="credit">Credit Note</option>
                                <option value="debit">Debit Note</option>
                                <option value="arbitration">Claim Arbitration</option>
                                <option value="currentmonth">Current Month Warranty</option>
                                <option value="compensation">Compensation Claim</option>
                                <option value="pr_approval">PR Approval</option>
                            </select>
                        </div>

                        <button onclick="exportToExcel()" class="export-btn" id="exportBtn">Export to Excel</button>
                    </div>
                </div>

                <div id="credit" class="tab-content active">
                    <div class="table-title">Warranty Credit Note by Division & Month</div>
                    <div class="table-wrapper">
                        <table class="data-table" id="creditTable"><thead></thead><tbody></tbody></table>
                    </div>
                </div>

                <div id="debit" class="tab-content">
                    <div class="table-title">Warranty Debit Note by Division & Month</div>
                    <div class="table-wrapper">
                        <table class="data-table" id="debitTable"><thead></thead><tbody></tbody></table>
                    </div>
                </div>

                <div id="arbitration" class="tab-content">
                    <div class="table-title">Pending Claim Arbitration by Division (Month-wise)</div>
                    <div class="table-wrapper">
                        <table class="data-table" id="arbitrationTable"><thead></thead><tbody></tbody></table>
                    </div>
                </div>

                <div id="currentmonth" class="tab-content">
                    <div class="table-title">Current Month Warranty - Pending Claims Summary</div>
                    <div class="table-wrapper">
                        <table class="data-table" id="currentMonthTable"><thead></thead><tbody></tbody></table>
                    </div>
                </div>

                <div id="compensation" class="tab-content">
                    <div class="table-title">Compensation Claim - Transit Claims Summary</div>
                    <div class="table-wrapper">
                        <table class="data-table" id="compensationTable"><thead></thead><tbody></tbody></table>
                    </div>
                </div>

                <div id="pr_approval" class="tab-content">
                    <div class="table-title">PR Approval - Claims Summary</div>
                    <div class="table-wrapper">
                        <table class="data-table" id="prApprovalTable"><thead></thead><tbody></tbody></table>
                    </div>
                </div>

            </div>
        </div>
    </div>

<script>
    let warrantyData = {};

    async function loadDashboard(){
        const spinner = document.getElementById('loadingSpinner');
        const tabs = document.getElementById('warrantyTabs');

        spinner.style.display = 'block';
        tabs.style.display = 'none';

        try{
            const response = await fetch('/api/warranty-data', {
                method:'GET',
                credentials:'include',
                headers:{ 'Content-Type':'application/json', 'Accept':'application/json' }
            });

            if(response.status === 401){
                alert('Session expired. Please login again.');
                window.location.href = '/login-page';
                return;
            }
            if(!response.ok){
                throw new Error('Failed to load warranty data: HTTP ' + response.status);
            }

            warrantyData = await response.json();

            displayTable('creditTable', warrantyData.credit, 0);
            displayTable('debitTable', warrantyData.debit, 0);
            displayTable('arbitrationTable', warrantyData.arbitration, 0);     // pending monthwise
            displayTable('currentMonthTable', warrantyData.currentMonth, 0);
            displayTable('compensationTable', warrantyData.compensation, 2);
            displayTable('prApprovalTable', warrantyData.prApproval, 2);

            loadDivisions();

            spinner.style.display = 'none';
            tabs.style.display = 'block';
        }catch(e){
            console.error(e);
            spinner.innerHTML = '<p style="color:red; padding:20px; text-align:center;">Error loading warranty data<br><br><button onclick="location.reload();" style="padding:10px 20px; background:#FF8C00; color:white; border:none; border-radius:6px; cursor:pointer; font-weight:600;">Refresh</button></p>';
        }
    }

    function displayTable(tableId, data, decimals){
        if(!data || data.length === 0) return;

        const table = document.getElementById(tableId);
        const headers = Object.keys(data[0]);

        const thead = table.querySelector('thead');
        thead.innerHTML = headers.map(h => '<th>' + h + '</th>').join('');

        const tbody = table.querySelector('tbody');
        tbody.innerHTML = data.map(row => {
            return '<tr>' + headers.map(h => {
                const v = row[h];
                if(typeof v === 'number'){
                    return '<td>' + v.toLocaleString('en-IN', {maximumFractionDigits: decimals}) + '</td>';
                }
                return '<td>' + (v ?? '') + '</td>';
            }).join('') + '</tr>';
        }).join('');
    }

    function switchTab(tabName){
        document.querySelectorAll('.tab-content').forEach(t => t.classList.remove('active'));
        document.querySelectorAll('.nav-link').forEach(b => b.classList.remove('active'));
        document.getElementById(tabName).classList.add('active');
        event.target.classList.add('active');
    }

    function loadDivisions(){
        const divisions = new Set();
        const type = document.getElementById('exportType').value;

        let dataSource = warrantyData.credit;
        if(type === 'debit') dataSource = warrantyData.debit;
        if(type === 'arbitration') dataSource = warrantyData.arbitration;
        if(type === 'currentmonth') dataSource = warrantyData.currentMonth;
        if(type === 'compensation') dataSource = warrantyData.compensation;
        if(type === 'pr_approval') dataSource = warrantyData.prApproval;

        if(dataSource && dataSource.length > 0){
            dataSource.forEach(r => {
                if(r.Division && r.Division !== 'Grand Total'){
                    divisions.add(r.Division);
                }
            });
        }

        const divisionSelect = document.getElementById('divisionFilter');
        const currentValue = divisionSelect.value;

        divisionSelect.innerHTML = '<option value="">-- Select Division --</option><option value="All">All Divisions</option>';

        Array.from(divisions).sort().forEach(div => {
            const opt = document.createElement('option');
            opt.value = div;
            opt.textContent = div;
            divisionSelect.appendChild(opt);
        });

        if(currentValue && divisionSelect.querySelector('option[value="' + currentValue + '"]')){
            divisionSelect.value = currentValue;
        }
    }

    document.getElementById('exportType')?.addEventListener('change', loadDivisions);

    async function exportToExcel(){
        const division = document.getElementById('divisionFilter').value;
        const type = document.getElementById('exportType').value;
        const exportBtn = document.getElementById('exportBtn');

        if(!division){
            alert('Please select a division');
            return;
        }

        exportBtn.disabled = true;
        exportBtn.textContent = 'Exporting...';

        try{
            const response = await fetch('/api/export-to-excel', {
                method:'POST',
                headers:{ 'Content-Type':'application/json' },
                credentials:'include',
                body: JSON.stringify({ division: division, type: type })
            });

            if(!response.ok){
                const err = await response.json().catch(() => ({detail:'Export failed'}));
                throw new Error(err.detail || 'Export failed');
            }

            const blob = await response.blob();
            const url = window.URL.createObjectURL(blob);
            const a = document.createElement('a');
            a.href = url;
            a.download = type + '_' + division + '_' + new Date().toISOString().split('T')[0] + '.xlsx';
            document.body.appendChild(a);
            a.click();
            window.URL.revokeObjectURL(url);
            document.body.removeChild(a);

            alert('Export completed successfully');
        }catch(e){
            alert('Export failed: ' + e.message);
        }finally{
            exportBtn.disabled = false;
            exportBtn.textContent = 'Export to Excel';
        }
    }

    window.onload = function(){ loadDashboard(); };
</script>

</body>
</html>
"""

# ==================== FASTAPI APP ====================

app = FastAPI()

# ==================== API ====================

@app.get("/api/captcha")
async def get_captcha():
    captcha_text, captcha_image = CaptchaGenerator.generate_captcha()
    return {"captcha": captcha_text, "image": captcha_image}

@app.get("/api/vehicle-images")
async def get_vehicle_images():
    return {"images": [{"name": img["name"], "data": img["data"]} for img in MAHINDRA_IMAGES]}

@app.post("/api/login")
async def api_login(request: Request):
    try:
        body = await request.json()
        user_id = body.get("user_id", "").strip()
        password = body.get("password", "")

        if not user_id or user_id not in USER_CREDENTIALS:
            raise HTTPException(status_code=401, detail="Invalid User ID")
        if USER_CREDENTIALS[user_id] != password:
            raise HTTPException(status_code=401, detail="Invalid Password")

        session_id = create_session(user_id)
        resp = JSONResponse(
            content={"success": True, "session_id": session_id, "user_id": user_id, "message": "Login successful"},
            status_code=200,
        )
        resp.set_cookie(
            key="session_id",
            value=session_id,
            httponly=True,
            max_age=28800,
            samesite="lax",
            path="/",
        )
        return resp
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))

@app.get("/api/warranty-data")
async def get_warranty_data(session_id: str = Cookie(None)):
    try:
        # If you want dashboard without login, comment this block.
        if not session_id or not verify_session(session_id):
            raise HTTPException(status_code=401, detail="Not authenticated")

        if WARRANTY_DATA["credit_df"] is None:
            return {
                "credit": [],
                "debit": [],
                "arbitration": [],
                "currentMonth": [],
                "compensation": [],
                "prApproval": [],
            }

        credit_records = WARRANTY_DATA["credit_df"].to_dict("records")
        debit_records = WARRANTY_DATA["debit_df"].to_dict("records")
        arbitration_records = WARRANTY_DATA["arbitration_df"].to_dict("records")  # pending monthwise
        current_month_records = WARRANTY_DATA["current_month_df"].to_dict("records") if WARRANTY_DATA["current_month_df"] is not None else []
        compensation_records = WARRANTYY_DATA_COMP = WARRANTY_DATA["compensation_df"].to_dict("records") if WARRANTY_DATA["compensation_df"] is not None else []
        pr_approval_records = WARRANTY_DATA["pr_approval_df"].to_dict("records") if WARRANTY_DATA["pr_approval_df"] is not None else []

        # replace NaN
        for records in [credit_records, debit_records, arbitration_records, current_month_records, compensation_records, pr_approval_records]:
            for record in records:
                for k in list(record.keys()):
                    if pd.isna(record[k]):
                        record[k] = 0

        return {
            "credit": credit_records,
            "debit": debit_records,
            "arbitration": arbitration_records,
            "currentMonth": current_month_records,
            "compensation": compensation_records,
            "prApproval": pr_approval_records,
        }

    except HTTPException:
        raise
    except Exception as e:
        import traceback
        print(f"  Unexpected error: {e}")
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))

# ==================== EXPORT ====================

@app.post("/api/export-to-excel")
async def export_to_excel(request: Request, session_id: str = Cookie(None)):
    try:
        # If you want export without login, comment this block.
        if not session_id or not verify_session(session_id):
            raise HTTPException(status_code=401, detail="Not authenticated")

        body = await request.json()
        selected_division = body.get("division", "All")
        export_type = body.get("type", "credit")

        if export_type not in ["credit", "debit", "arbitration", "currentmonth", "compensation", "pr_approval"]:
            raise HTTPException(status_code=400, detail="Invalid export type")

        if export_type == "currentmonth":
            return await export_current_month_warranty(selected_division)

        if export_type == "compensation":
            return await export_compensation_claim(selected_division)

        if export_type == "pr_approval":
            return await export_pr_approval(selected_division)

        if export_type == "credit":
            df = WARRANTY_DATA["credit_df"]
        elif export_type == "debit":
            df = WARRANTY_DATA["debit_df"]
        else:
            df = WARRANTY_DATA["arbitration_df"]  # pending monthwise summary

        if df is None or df.empty:
            raise HTTPException(status_code=500, detail="No data available for export")

        dealer_mapping = {
            "AMRAVATI": "AMT",
            "CHAUFULA_SZZ": "CHA",
            "CHIKHALI": "CHI",
            "KOLHAPUR_WS": "KOL",
            "NAGPUR_KAMPTHEE ROAD": "HO",
            "NAGPUR_WARDHAMAN NGR": "CITY",
            "SHIKRAPUR_SZS": "SHI",
            "WAGHOLI": "WAG",
            "YAVATMAL": "YAT",
            "NAGPUR_WARDHAMAN NGR_CQ": "CQ",
        }
        reverse_mapping = {v: k for k, v in dealer_mapping.items()}

        if selected_division not in ["All", "Grand Total"]:
            df_export = df[df["Division"] == selected_division].copy()
            gt = df[df["Division"] == "Grand Total"]
            if not gt.empty:
                df_export = pd.concat([df_export, gt], ignore_index=True)
        else:
            df_export = df.copy()

        wb = Workbook()
        header_fill = PatternFill(start_color="FF8C00", end_color="FF8C00", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))

        # Sheet 1 Summary
        ws1 = wb.active
        ws1.title = f"{selected_division} - {export_type.capitalize()}" if selected_division not in ["All", "Grand Total"] else export_type.capitalize()

        for col_idx, column in enumerate(df_export.columns, 1):
            c = ws1.cell(row=1, column=col_idx, value=column)
            c.fill = header_fill
            c.font = header_font
            c.border = border
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        for row_idx, row in enumerate(df_export.itertuples(index=False), 2):
            for col_idx, value in enumerate(row, 1):
                cell = ws1.cell(row=row_idx, column=col_idx)
                if isinstance(value, (int, float, np.integer, np.floating)):
                    cell.value = float(value)
                    cell.number_format = "#,##0.00"
                    cell.alignment = Alignment(horizontal="right", vertical="center")
                else:
                    cell.value = str(value) if value is not None else ""
                    cell.alignment = Alignment(horizontal="left", vertical="center")
                cell.border = border

        for col_idx, column in enumerate(df_export.columns, 1):
            max_length = min(
                max(df_export[column].astype(str).map(len).max(), len(str(column))) + 2,
                30,
            )
            ws1.column_dimensions[get_column_letter(col_idx)].width = max_length

        # Sheet 2 Detailed source (only for single division)
        if selected_division not in ["All", "Grand Total"]:
            ws2 = wb.create_sheet()
            ws2.title = f"{selected_division} - Detailed Data"
            dealer_location = reverse_mapping.get(selected_division)

            if dealer_location and WARRANTY_DATA["source_df"] is not None:
                source_df = WARRANTY_DATA["source_df"].copy()
                detail_df = source_df[source_df["Dealer Location"] == dealer_location].copy()

                def is_empty_or_hyphen(value):
                    if pd.isna(value):
                        return True
                    v = str(value).strip()
                    if v == "" or v == "-" or v.upper() == "NAN":
                        return True
                    return False

                required_columns = [
                    "Fiscal Month",
                    "Dealer Location",
                    "Claim arbitration ID",
                    "Claim Invoice Date",
                    "Claim No",
                    "Claim Date",
                    "Chassis No",
                    "Ro Id",
                    "Claim Type",
                ]

                # Export logic per type
                if export_type == "credit":
                    detail_df = detail_df[detail_df["Credit Note Amount"] > 0].copy()
                    detail_df = detail_df[detail_df["Claim arbitration ID"].apply(is_empty_or_hyphen)].copy()
                    required_columns += ["Total Claim Amount", "Credit Note Amount"]

                elif export_type == "debit":
                    detail_df = detail_df[detail_df["Debit Note Amount"] > 0].copy()
                    required_columns += ["Total Claim Amount", "Debit Note Amount"]

                else:
                    # arbitration tab is pending monthwise -> detailed should also be PENDING records
                    detail_df = detail_df[detail_df["Debit Note Amount"] > 0].copy()
                    detail_df = detail_df[detail_df["Claim arbitration ID"].apply(is_empty_or_hyphen)].copy()
                    required_columns += ["Total Claim Amount", "Credit Note Amount", "Debit Note Amount"]

                available_columns = [c for c in required_columns if c in detail_df.columns]
                detail_df = detail_df[available_columns].copy()

                if "Claim No" in detail_df.columns:
                    def format_claim_no(x):
                        if pd.isna(x) or str(x).strip() == "":
                            return ""
                        try:
                            return str(int(float(x)))
                        except Exception:
                            return str(x).strip()
                    detail_df["Claim No"] = detail_df["Claim No"].apply(format_claim_no)

                if "Ro Id" in detail_df.columns:
                    def format_ro_id(x):
                        if pd.isna(x) or str(x).strip() == "":
                            return ""
                        try:
                            return f"RO{str(int(float(x)))}"
                        except Exception:
                            s = str(x).strip()
                            return s if s.startswith("RO") else f"RO{s}"
                    detail_df["Ro Id"] = detail_df["Ro Id"].apply(format_ro_id)

                # Rename for clarity when export_type == arbitration (pending)
                if export_type == "arbitration" and "Debit Note Amount" in detail_df.columns:
                    detail_df = detail_df.rename(columns={"Debit Note Amount": "Pending Arbitration Amount"})

                # Sort by fiscal month order
                month_order = ["Apr","May","Jun","Jul","Aug","Sep","Oct","Nov","Dec","Jan","Feb","Mar"]
                if "Fiscal Month" in detail_df.columns:
                    detail_df["Month"] = detail_df["Fiscal Month"].astype(str).str.strip().str[:3]
                    detail_df["Month_Order"] = detail_df["Month"].apply(lambda x: month_order.index(x) if x in month_order else 999)
                    detail_df = detail_df.sort_values("Month_Order").drop(["Month","Month_Order"], axis=1)

                for col_idx, column in enumerate(detail_df.columns, 1):
                    c = ws2.cell(row=1, column=col_idx, value=column)
                    c.fill = header_fill
                    c.font = header_font
                    c.border = border
                    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

                for row_idx, row in enumerate(detail_df.itertuples(index=False), 2):
                    for col_idx, value in enumerate(row, 1):
                        cell = ws2.cell(row=row_idx, column=col_idx)
                        col_name = detail_df.columns[col_idx - 1]

                        if col_name in ["Claim No", "Ro Id"]:
                            cell.value = str(value) if value is not None else ""
                            cell.alignment = Alignment(horizontal="left", vertical="center")
                        elif isinstance(value, (int, float, np.integer, np.floating)):
                            cell.value = float(value)
                            cell.number_format = "#,##0.00"
                            cell.alignment = Alignment(horizontal="right", vertical="center")
                        elif isinstance(value, (datetime, pd.Timestamp)):
                            cell.value = value
                            cell.number_format = "mm-dd-yyyy"
                            cell.alignment = Alignment(horizontal="center", vertical="center")
                        else:
                            cell.value = str(value) if not pd.isna(value) else ""
                            cell.alignment = Alignment(horizontal="left", vertical="center")
                        cell.border = border

                for col_idx, column in enumerate(detail_df.columns, 1):
                    max_length = min(
                        max(detail_df[column].astype(str).map(len).max(), len(str(column))) + 2,
                        35,
                    )
                    ws2.column_dimensions[get_column_letter(col_idx)].width = max_length

        out = io.BytesIO()
        wb.save(out)
        out.seek(0)

        filename = f"{selected_division}_{export_type}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        return StreamingResponse(
            iter([out.getvalue()]),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"},
        )

    except HTTPException:
        raise
    except Exception as e:
        import traceback
        print(f"  Export error: {e}")
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"Export error: {str(e)}")

async def export_current_month_warranty(selected_division: str):
    summary_df = WARRANTY_DATA["current_month_df"]
    source_df = WARRANTY_DATA["current_month_source_df"]
    if summary_df is None or summary_df.empty:
        raise HTTPException(status_code=500, detail="No current month warranty data available")

    if selected_division not in ["All", "Grand Total"]:
        df_export = summary_df[summary_df["Division"] == selected_division].copy()
        gt = summary_df[summary_df["Division"] == "Grand Total"]
        if not gt.empty:
            df_export = pd.concat([df_export, gt], ignore_index=True)
    else:
        df_export = summary_df.copy()

    wb = Workbook()
    header_fill = PatternFill(start_color="FF8C00", end_color="FF8C00", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))

    ws1 = wb.active
    ws1.title = f"{selected_division} - Summary" if selected_division not in ["All", "Grand Total"] else "Current Month Summary"

    for col_idx, column in enumerate(df_export.columns, 1):
        c = ws1.cell(row=1, column=col_idx, value=column)
        c.fill = header_fill
        c.font = header_font
        c.border = border
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for row_idx, row in enumerate(df_export.itertuples(index=False), 2):
        for col_idx, value in enumerate(row, 1):
            cell = ws1.cell(row=row_idx, column=col_idx)
            if isinstance(value, (int, float, np.integer, np.floating)):
                cell.value = float(value)
                cell.number_format = "#,##0"
                cell.alignment = Alignment(horizontal="right", vertical="center")
            else:
                cell.value = str(value) if value is not None else ""
                cell.alignment = Alignment(horizontal="left", vertical="center")
            cell.border = border

    for col_idx, column in enumerate(df_export.columns, 1):
        max_length = min(max(df_export[column].astype(str).map(len).max(), len(str(column))) + 2, 30)
        ws1.column_dimensions[get_column_letter(col_idx)].width = max_length

    # Sheet 2 Spares
    if source_df is not None and not source_df.empty:
        spares_df = source_df.copy()
        if selected_division not in ["All", "Grand Total"]:
            spares_df = spares_df[spares_df["Division"] == selected_division].copy()
        spares_df = spares_df[spares_df["Pending Claims Spares"].notna()].copy()

        if not spares_df.empty:
            ws2 = wb.create_sheet()
            ws2.title = f"{selected_division} - Spares" if selected_division not in ["All", "Grand Total"] else "Pending Spares Claims"

            for col_idx, column in enumerate(spares_df.columns, 1):
                c = ws2.cell(row=1, column=col_idx, value=column)
                c.fill = header_fill
                c.font = header_font
                c.border = border
                c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

            for row_idx, row in enumerate(spares_df.itertuples(index=False), 2):
                for col_idx, value in enumerate(row, 1):
                    cell = ws2.cell(row=row_idx, column=col_idx)
                    if isinstance(value, (int, float, np.integer, np.floating)):
                        cell.value = float(value)
                        cell.number_format = "#,##0.00"
                        cell.alignment = Alignment(horizontal="right", vertical="center")
                    elif isinstance(value, (datetime, pd.Timestamp)):
                        cell.value = value
                        cell.number_format = "mm-dd-yyyy"
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                    else:
                        cell.value = str(value) if not pd.isna(value) else ""
                        cell.alignment = Alignment(horizontal="left", vertical="center")
                    cell.border = border

            for col_idx, column in enumerate(spares_df.columns, 1):
                max_length = min(max(spares_df[column].astype(str).map(len).max(), len(str(column))) + 2, 35)
                ws2.column_dimensions[get_column_letter(col_idx)].width = max_length

    # Sheet 3 Labour
    if source_df is not None and not source_df.empty:
        labour_df = source_df.copy()
        if selected_division not in ["All", "Grand Total"]:
            labour_df = labour_df[labour_df["Division"] == selected_division].copy()
        labour_df = labour_df[labour_df["Pending Claims Labour"].notna()].copy()

        if not labour_df.empty:
            ws3 = wb.create_sheet()
            ws3.title = f"{selected_division} - Labour" if selected_division not in ["All", "Grand Total"] else "Pending Labour Claims"

            for col_idx, column in enumerate(labour_df.columns, 1):
                c = ws3.cell(row=1, column=col_idx, value=column)
                c.fill = header_fill
                c.font = header_font
                c.border = border
                c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

            for row_idx, row in enumerate(labour_df.itertuples(index=False), 2):
                for col_idx, value in enumerate(row, 1):
                    cell = ws3.cell(row=row_idx, column=col_idx)
                    if isinstance(value, (int, float, np.integer, np.floating)):
                        cell.value = float(value)
                        cell.number_format = "#,##0.00"
                        cell.alignment = Alignment(horizontal="right", vertical="center")
                    elif isinstance(value, (datetime, pd.Timestamp)):
                        cell.value = value
                        cell.number_format = "mm-dd-yyyy"
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                    else:
                        cell.value = str(value) if not pd.isna(value) else ""
                        cell.alignment = Alignment(horizontal="left", vertical="center")
                    cell.border = border

            for col_idx, column in enumerate(labour_df.columns, 1):
                max_length = min(max(labour_df[column].astype(str).map(len).max(), len(str(column))) + 2, 35)
                ws3.column_dimensions[get_column_letter(col_idx)].width = max_length

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    filename = f"{selected_division}_CurrentMonthWarranty_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return StreamingResponse(
        iter([out.getvalue()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )

async def export_compensation_claim(selected_division: str):
    summary_df = WARRANTY_DATA["compensation_df"]
    source_df = WARRANTY_DATA["compensation_source_df"]
    if summary_df is None or summary_df.empty:
        raise HTTPException(status_code=500, detail="No compensation claim data available")

    if selected_division not in ["All", "Grand Total"]:
        df_export = summary_df[summary_df["Division"] == selected_division].copy()
        gt = summary_df[summary_df["Division"] == "Grand Total"]
        if not gt.empty:
            df_export = pd.concat([df_export, gt], ignore_index=True)
    else:
        df_export = summary_df.copy()

    wb = Workbook()
    header_fill = PatternFill(start_color="FF8C00", end_color="FF8C00", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))

    ws1 = wb.active
    ws1.title = f"{selected_division} - Summary" if selected_division not in ["All", "Grand Total"] else "Compensation Summary"

    for col_idx, column in enumerate(df_export.columns, 1):
        c = ws1.cell(row=1, column=col_idx, value=column)
        c.fill = header_fill
        c.font = header_font
        c.border = border
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for row_idx, row in enumerate(df_export.itertuples(index=False), 2):
        for col_idx, value in enumerate(row, 1):
            cell = ws1.cell(row=row_idx, column=col_idx)
            if isinstance(value, (int, float, np.integer, np.floating)):
                cell.value = float(value)
                cell.number_format = "#,##0.00"
                cell.alignment = Alignment(horizontal="right", vertical="center")
            else:
                cell.value = str(value) if value is not None else ""
                cell.alignment = Alignment(horizontal="left", vertical="center")
            cell.border = border

    for col_idx, column in enumerate(df_export.columns, 1):
        max_length = min(max(df_export[column].astype(str).map(len).max(), len(str(column))) + 2, 30)
        ws1.column_dimensions[get_column_letter(col_idx)].width = max_length

    if source_df is not None and not source_df.empty:
        detail_df = source_df.copy()
        if selected_division not in ["All", "Grand Total"]:
            detail_df = detail_df[detail_df["Division"] == selected_division].copy()

        if not detail_df.empty:
            ws2 = wb.create_sheet()
            ws2.title = f"{selected_division} - Details" if selected_division not in ["All", "Grand Total"] else "Compensation Details"

            for col_idx, column in enumerate(detail_df.columns, 1):
                c = ws2.cell(row=1, column=col_idx, value=column)
                c.fill = header_fill
                c.font = header_font
                c.border = border
                c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

            for row_idx, row in enumerate(detail_df.itertuples(index=False), 2):
                for col_idx, value in enumerate(row, 1):
                    cell = ws2.cell(row=row_idx, column=col_idx)
                    if isinstance(value, (int, float, np.integer, np.floating)):
                        cell.value = float(value)
                        cell.number_format = "#,##0.00"
                        cell.alignment = Alignment(horizontal="right", vertical="center")
                    elif isinstance(value, (datetime, pd.Timestamp)):
                        cell.value = value
                        cell.number_format = "mm-dd-yyyy"
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                    else:
                        cell.value = str(value) if not pd.isna(value) else ""
                        cell.alignment = Alignment(horizontal="left", vertical="center")
                    cell.border = border

            for col_idx, column in enumerate(detail_df.columns, 1):
                max_length = min(max(detail_df[column].astype(str).map(len).max(), len(str(column))) + 2, 35)
                ws2.column_dimensions[get_column_letter(col_idx)].width = max_length

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    filename = f"{selected_division}_CompensationClaim_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return StreamingResponse(
        iter([out.getvalue()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )

async def export_pr_approval(selected_division: str):
    summary_df = WARRANTY_DATA["pr_approval_df"]
    source_df = WARRANTY_DATA["pr_approval_source_df"]
    if summary_df is None or summary_df.empty:
        raise HTTPException(status_code=500, detail="No PR Approval data available")

    if selected_division not in ["All", "Grand Total"]:
        df_export = summary_df[summary_df["Division"] == selected_division].copy()
        gt = summary_df[summary_df["Division"] == "Grand Total"]
        if not gt.empty:
            df_export = pd.concat([df_export, gt], ignore_index=True)
    else:
        df_export = summary_df.copy()

    wb = Workbook()
    header_fill = PatternFill(start_color="FF8C00", end_color="FF8C00", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))

    ws1 = wb.active
    ws1.title = f"{selected_division} - Summary" if selected_division not in ["All", "Grand Total"] else "PR Approval Summary"

    for col_idx, column in enumerate(df_export.columns, 1):
        c = ws1.cell(row=1, column=col_idx, value=column)
        c.fill = header_fill
        c.font = header_font
        c.border = border
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for row_idx, row in enumerate(df_export.itertuples(index=False), 2):
        for col_idx, value in enumerate(row, 1):
            cell = ws1.cell(row=row_idx, column=col_idx)
            if isinstance(value, (int, float, np.integer, np.floating)):
                cell.value = float(value)
                cell.number_format = "#,##0.00"
                cell.alignment = Alignment(horizontal="right", vertical="center")
            else:
                cell.value = str(value) if value is not None else ""
                cell.alignment = Alignment(horizontal="left", vertical="center")
            cell.border = border

    for col_idx, column in enumerate(df_export.columns, 1):
        max_length = min(max(df_export[column].astype(str).map(len).max(), len(str(column))) + 2, 30)
        ws1.column_dimensions[get_column_letter(col_idx)].width = max_length

    if source_df is not None and not source_df.empty:
        detail_df = source_df.copy()
        if selected_division not in ["All", "Grand Total"]:
            detail_df = detail_df[detail_df["Division"] == selected_division].copy()

        if not detail_df.empty:
            ws2 = wb.create_sheet()
            ws2.title = f"{selected_division} - Details" if selected_division not in ["All", "Grand Total"] else "PR Approval Details"

            for col_idx, column in enumerate(detail_df.columns, 1):
                c = ws2.cell(row=1, column=col_idx, value=column)
                c.fill = header_fill
                c.font = header_font
                c.border = border
                c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

            for row_idx, row in enumerate(detail_df.itertuples(index=False), 2):
                for col_idx, value in enumerate(row, 1):
                    cell = ws2.cell(row=row_idx, column=col_idx)
                    if isinstance(value, (int, float, np.integer, np.floating)):
                        cell.value = float(value)
                        cell.number_format = "#,##0.00"
                        cell.alignment = Alignment(horizontal="right", vertical="center")
                    elif isinstance(value, (datetime, pd.Timestamp)):
                        cell.value = value
                        cell.number_format = "mm-dd-yyyy"
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                    else:
                        cell.value = str(value) if not pd.isna(value) else ""
                        cell.alignment = Alignment(horizontal="left", vertical="center")
                    cell.border = border

            for col_idx, column in enumerate(detail_df.columns, 1):
                max_length = min(max(detail_df[column].astype(str).map(len).max(), len(str(column))) + 2, 35)
                ws2.column_dimensions[get_column_letter(col_idx)].width = max_length

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    filename = f"{selected_division}_PrApproval_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return StreamingResponse(
        iter([out.getvalue()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )

# ==================== ROUTES ====================

@app.get("/login-page")
async def login_page():
    return HTMLResponse(content=LOGIN_PAGE)

@app.get("/dashboard")
async def dashboard():
    return HTMLResponse(content=DASHBOARD_HTML)

@app.get("/")
async def root():
    return HTMLResponse(content=DASHBOARD_HTML)

# ==================== STARTUP LOAD ====================

print("\n" + "=" * 100)
print("STARTING WARRANTY MANAGEMENT SYSTEM")
print("=" * 100)

print("\nProcessing warranty data...")
WARRANTY_DATA["credit_df"], WARRANTY_DATA["debit_df"], WARRANTY_DATA["arbitration_df"], WARRANTY_DATA["source_df"] = process_warranty_data()

print("\nProcessing current month warranty data...")
WARRANTY_DATA["current_month_df"], WARRANTY_DATA["current_month_source_df"] = process_current_month_warranty()

print("\nProcessing compensation claim data...")
WARRANTY_DATA["compensation_df"], WARRANTY_DATA["compensation_source_df"] = process_compensation_claim()

print("\nProcessing PR Approval data...")
WARRANTY_DATA["pr_approval_df"], WARRANTY_DATA["pr_approval_source_df"] = process_pr_approval()

if __name__ == "__main__":
    hostname = socket.gethostname()
    try:
        local_ip = socket.gethostbyname(hostname)
    except Exception:
        local_ip = "127.0.0.1"

    port = int(os.getenv("PORT", "8001"))

    print("\n" + "=" * 100)
    print("SERVER READY - Warranty Dashboard")
    print("=" * 100)
    print(f"PORT: {port}")
    print(f"Login URL: http://localhost:{port}/login-page")
    print(f"Dashboard URL: http://localhost:{port}/dashboard")
    print(f"Network URL: http://{local_ip}:{port}/login-page")
    print("=" * 100 + "\n")

    uvicorn.run(app, host="0.0.0.0", port=port)
